"""Export assessment plans to PDF, Excel, and JSON."""

import json
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from fpdf import FPDF
from openpyxl import Workbook

from app.config import DOCUMENTS_DIR, settings
from app.models import DiscoverySession
from app.schemas import SessionState
from app.services.assessment_plan import build_assessment_plan


@dataclass
class ExportResult:
    pdf_path: str
    excel_path: str
    json_path: str
    stem: str


def sanitize_filename_part(value: str, max_len: int = 40) -> str:
    cleaned = re.sub(r"[^\w\s-]", "", value or "").strip().replace(" ", "_")
    if not cleaned:
        return "unknown"
    return cleaned[:max_len]


def build_filename_stem(record: DiscoverySession, state: SessionState) -> str:
    company = sanitize_filename_part(state.company.company_name or record.company_name)
    products = sanitize_filename_part(state.company.products or record.products)
    if company == "unknown" and products == "unknown":
        return f"session_{record.id}"
    return f"{company}_{products}"


def _export_json(path: Path, record: DiscoverySession, state: SessionState, plan: dict) -> None:
    payload = {
        "access_key": record.access_key,
        "company_name": state.company.company_name,
        "products": state.company.products,
        "company": state.company.model_dump(),
        "objectives": state.objectives.model_dump(),
        "applications": [a.model_dump() for a in state.applications],
        "answers": [a.model_dump() for a in state.answers],
        "plan": plan,
        "exported_at": datetime.utcnow().isoformat(),
    }
    path.write_text(json.dumps(payload, indent=2, default=str), encoding="utf-8")


def _export_excel(path: Path, record: DiscoverySession, state: SessionState, plan: dict) -> None:
    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    summary.append(["Company", state.company.company_name])
    summary.append(["Products", state.company.products])
    summary.append(["Access Key", record.access_key])
    summary.append(["Approved", state.plan_approved])
    summary.append(["Coverage %", plan.get("coverage_summary", {}).get("percent_complete")])

    company_sheet = wb.create_sheet("Company")
    for key, val in state.company.model_dump().items():
        company_sheet.append([key, val])

    objectives_sheet = wb.create_sheet("Objectives")
    for key, val in state.objectives.model_dump().items():
        objectives_sheet.append([key, str(val)])

    apps_sheet = wb.create_sheet("Applications")
    apps_sheet.append(
        ["name", "criticality", "exposure", "data_classes", "repositories", "owners"]
    )
    for app in state.applications:
        apps_sheet.append(
            [
                app.name,
                app.criticality,
                app.exposure,
                app.data_classes,
                app.repositories,
                f"{app.business_owner}/{app.technical_owner}",
            ]
        )

    coverage_sheet = wb.create_sheet("Coverage")
    for key, val in plan.get("coverage_summary", {}).items():
        coverage_sheet.append([key, val])

    missing_sheet = wb.create_sheet("MissingInfo")
    missing_sheet.append(["question_id", "applies_to", "notes"])
    for item in plan.get("missing_information_register", []):
        missing_sheet.append([item.get("question_id"), item.get("applies_to"), item.get("notes")])

    plan_sheet = wb.create_sheet("ApplicationsPlan")
    for app_plan in plan.get("applications", []):
        plan_sheet.append(["application", app_plan.get("application")])
        plan_sheet.append(["criticality", app_plan.get("criticality")])
        plan_sheet.append(["exposure", app_plan.get("exposure")])
        plan_sheet.append(["scans", ", ".join(app_plan.get("required_scans", []))])
        plan_sheet.append([])

    wb.save(path)


def _safe_text(text: str) -> str:
    return (text or "").encode("latin-1", errors="replace").decode("latin-1")


def _export_pdf(
    path: Path,
    record: DiscoverySession,
    state: SessionState,
    plan: dict,
    base_url: str,
) -> None:
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 10, _safe_text("Technical Debt Assessment Plan"), ln=True)
    pdf.set_font("Helvetica", "", 11)
    pdf.cell(0, 8, _safe_text(f"Company: {state.company.company_name}"), ln=True)
    pdf.cell(0, 8, _safe_text(f"Products: {state.company.products}"), ln=True)
    pdf.cell(0, 8, _safe_text(f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"), ln=True)
    pdf.ln(4)

    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 8, "Objectives", ln=True)
    pdf.set_font("Helvetica", "", 11)
    pdf.multi_cell(pdf.epw, 6, _safe_text(f"Primary goal: {state.objectives.primary_goal}"))
    pdf.multi_cell(pdf.epw, 6, _safe_text(f"Audiences: {', '.join(state.objectives.report_audiences)}"))
    pdf.multi_cell(pdf.epw, 6, _safe_text(f"Release blockers: {', '.join(state.objectives.release_blockers)}"))
    pdf.ln(4)

    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 8, "Applications", ln=True)
    pdf.set_font("Helvetica", "", 11)
    for app in state.applications:
        pdf.multi_cell(pdf.epw, 6, _safe_text(f"- {app.name}: {app.criticality}, {app.exposure}, {app.data_classes}"))

    pdf.ln(4)
    cov = plan.get("coverage_summary", {})
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 8, "Coverage", ln=True)
    pdf.set_font("Helvetica", "", 11)
    pdf.multi_cell(
        pdf.epw,
        6,
        _safe_text(
            f"Completed: {cov.get('completed')} | Missing: {cov.get('missing')} | "
            f"Unknown: {cov.get('unknown')} | {cov.get('percent_complete')}% resolved"
        ),
    )

    pdf.ln(8)
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 8, "Retrieval key", ln=True)
    pdf.set_font("Helvetica", "", 10)
    pdf.multi_cell(pdf.epw, 6, _safe_text("Save this key to retrieve and update your submission later:"))
    pdf.set_font("Helvetica", "B", 9)
    pdf.multi_cell(pdf.epw, 5, _safe_text(record.access_key))
    pdf.set_font("Helvetica", "", 10)
    pdf.multi_cell(pdf.epw, 6, _safe_text(f"Visit: {base_url.rstrip('/')}/retrieve"))

    pdf.output(path)


def export_all(
    record: DiscoverySession,
    state: SessionState,
    base_url: str | None = None,
    documents_root: Path | None = None,
) -> ExportResult:
    root = documents_root or DOCUMENTS_DIR
    session_dir = root / str(record.id)
    session_dir.mkdir(parents=True, exist_ok=True)

    stem = build_filename_stem(record, state)
    plan = build_assessment_plan(state)
    base = base_url or settings.app_base_url

    json_path = session_dir / f"{stem}.json"
    excel_path = session_dir / f"{stem}.xlsx"
    pdf_path = session_dir / f"{stem}.pdf"

    _export_json(json_path, record, state, plan)
    _export_excel(excel_path, record, state, plan)
    _export_pdf(pdf_path, record, state, plan, base)

    rel_json = f"documents/{record.id}/{stem}.json"
    rel_excel = f"documents/{record.id}/{stem}.xlsx"
    rel_pdf = f"documents/{record.id}/{stem}.pdf"

    return ExportResult(
        pdf_path=rel_pdf,
        excel_path=rel_excel,
        json_path=rel_json,
        stem=stem,
    )
